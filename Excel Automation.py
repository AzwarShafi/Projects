import openpyxl as xl
#loads the excel sheet in python
wb = xl.load_workbook('/Users/azwarshafi/Library/CloudStorage/OneDrive-Personal/Desktop/Python_/Projects/Python Automation File.xlsx')
#selects the specific sheet
ws = wb['Sheet1']
#changes value and prints that value of a specific cell in sheet
print(ws['A2'].value)
#---------------------
for row in range(2,5):
    result = ws.cell(row,2)
    discounted_price = result.value / 10
    discounted_cell = ws.cell(row,3)
    discounted_cell.value = discounted_price

wb.save('automated sheet.xlsx')

